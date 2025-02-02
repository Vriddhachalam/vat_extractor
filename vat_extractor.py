# gemini-04-12-24

import os, re, json
import google.generativeai as genai
from datetime import datetime
from copy import copy
from openpyxl import load_workbook, utils
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def table_updater(sheet_name,table_name,data):
    sheet = workbook[sheet_name]  # or workbook['SheetName']
    table = sheet.tables[table_name]
    table_range = table.ref  # Get the range of the table (e.g., 'A1:C10')
    start_row = int(table_range.split(':')[0][1:])  # Get the starting row
    start_col_letter = table_range.split(':')[0][0]  # Get the starting column letter
    start_column = utils.cell.column_index_from_string(start_col_letter)  # Convert letter to index

    next_row = sheet.max_row + 1  # Start checking from the last used row

    while next_row <= sheet.max_row and any(sheet.cell(row=next_row, column=col).value is not None for col in range(start_column, start_column + len(data))):
        next_row += 1

    for idx, value in enumerate(data.values()):
        sheet.cell(row=next_row, column=start_column + idx, value=value)

    last_row = next_row - 1  # The row to copy formatting from
    if last_row >= start_row:
        source_row = sheet[last_row]
        target_row = sheet[next_row]
        copy_row_formatting(source_row, target_row)

    new_table_range = f"{table_range.split(':')[0]}:{get_column_letter(start_column + len(data) - 1)}{next_row}"
    table.ref = new_table_range  # Update the table reference

    workbook.save(file_path)    

def list_files_in_directory(directory):
    all_files = []
    for dirpath, dirnames, filenames in os.walk(directory):
        for filename in filenames:
            all_files.append(os.path.join(dirpath, filename))  # Create the full path
    return all_files

def copy_row_formatting(source_row, target_row):
    for source_cell in source_row:
        target_cell = target_row[source_cell.column - 1]  # Adjust for 1-based indexing
        # Copy styles individually
        target_cell.font = Font(name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                vertAlign=source_cell.font.vertAlign,
                                underline=source_cell.font.underline,
                                strike=source_cell.font.strike,
                                color=source_cell.font.color)
        
        target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                           vertical=source_cell.alignment.vertical,
                                           text_rotation=source_cell.alignment.text_rotation,
                                           wrap_text=source_cell.alignment.wrap_text,
                                           shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                           indent=source_cell.alignment.indent)
        
        target_cell.fill = copy(source_cell.fill)  # Use copy for fill styles
        target_cell.border = copy(source_cell.border)  # Copy border styles
        target_cell.number_format = copy(source_cell.number_format)  # Direct assignment is fine
        target_cell.protection = copy(source_cell.protection)  # Copy protection
        target_cell.hyperlink = copy(source_cell.hyperlink)  # Direct assignment is 

def zero_if_null(value):
    return 0 if value is None else value

with open('./config.txt', 'r') as file:
    # Initialize variables
    API_KEY = None
    MODEL = None
    # Read each line and process
    for line in file:
        # Remove any surrounding whitespace or newline characters
        line = line.strip()
        # Split the line into key-value pairs
        if line.startswith('api_key'):
            API_KEY = line.split('=')[1].strip()  # Get the API key
        elif line.startswith('model'):
            MODEL = line.split('=')[1].strip()  # Get the serial ID    

print('******************************WELCOME TO VAT EXTRACTOR********************************')
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel(MODEL)

file_path = './vat_purchase_invoices.xlsx'
workbook = load_workbook(file_path)
directory = './invoices/'
files = list_files_in_directory(directory)
print("FILES_COUNT:", len(files))
failed_files = {}

ws = workbook['invoices']  # or workbook['SheetName']
table = ws.tables['Table1']
# Get the header row for the table to find the "File" column
header_row = ws[table.ref.split(':')[0][1:]]
# Find the index of the "File" column
file_column_index = None
for idx, cell in enumerate(header_row):
    if cell.value == "File":
        file_column_index = idx + 1  # openpyxl is 1-indexed
        break
if file_column_index is None:
    raise ValueError("The column 'File' was not found in the table.")
# Now extract the values from the 'File' column into a list
existing_files = []
# Loop through the rows of the table and extract values from the 'File' column
for row in ws[table.ref]:
    existing_files.append(row[file_column_index - 1].value)  # Adjust for zero-indexing in Python

for filename in files:

    print(f"\n--- FILE: {files.index(filename)+1} OF {len(files)}, {filename} ---")
    if filename not in existing_files:
        invoice_type, invoice_meant_for, alter_point = ('Inward', 'from','to') if './invoices/inward' in filename else ('Outward', 'to','from')

        prompt = f'''Return the following from the uploaded invoice (json bstring) (DO NOT OVERFIT try to read properly and extract) :
                                                vat percentage as vat% (float, DO NOT ROUND OFF),
                                                the vat or tva as vat_amount (float, DO NOT ROUND OFF),
                                                the net total excluding vat as total_amount (float, DO NOT ROUND OFF),
                                                The company_name from '{invoice_meant_for}' of the invoice:  return as company_name (string).
                                                The alter_company_name from '{alter_point}' of the invoice:  return as alter_company_name (string).
                                                invoice date as invoice_date in (string force to numerical DD/MM/YYYY format ),
                                                invoice no as invoice_no in (string)
                                                if different vat% or vat_amount exists  add them as separate keys  
                                            '''
        if filename.endswith('.pdf'):
            try:
                sample_doc = genai.upload_file(os.path.join(filename))
                response = model.generate_content([prompt, sample_doc])

                data_string = response.text
                json_match = re.search(r'\{.*?\}', data_string, re.DOTALL)

                if json_match:
                    json_data = json_match.group(0)  # Extract the JSON string
                    row = json.loads(json_data)
                
                keywords = ['lining', 'enquire', 'barry','reilly']  # Add more keywords as needed
                # Check if any of the keywords is in the company_name (case-insensitive)
                row['alter_company_name'] if any(keyword in row['company_name'].lower() for keyword in keywords) else row['company_name']

                to_append = {
                    'Timestamp': str(datetime.now()),
                    'File': str(filename),
                    'Invoice Type': invoice_type,
                    'Company Name': row['alter_company_name'] if any(keyword in row['company_name'].lower() for keyword in keywords) else row['company_name'],
                    'Invoice/Credit Note#': row['invoice_no'],
                    'Date': row['invoice_date'],
                    'Amount (Excl Vat)': row['total_amount'],
                    'Purchases @ Zero Vat': row['total_amount'] if str(row['vat%'])[:3] in ('0.0', 'Non') else '-',
                    'VAT @ 13.5%': row['vat_amount'] if '13.5' in str(row['vat%']) else '-',
                    'VAT @ 21%': row['vat_amount'] if '21' in str(row['vat%']) else '-',
                    'VAT @ 23%': row['vat_amount'] if '23' in str(row['vat%']) else '-',
                    'Total Vat': row['vat_amount'],
                    'Total': float(zero_if_null(row['total_amount'])) + float(zero_if_null(row['vat_amount'])),
                    'Notes': '-' if str(row['vat%'])[:3] in ('0.0', 'Non') else row['vat_amount']
                }

                table_updater('invoices','Table1',to_append)
                print("******************SUCCESS******************")

            except Exception as err:
                print("******************FAILED******************")
                print('EXCEPTION:\n', err, '\n', 'FILE_NAME:', filename)
                failed_files['Datetime'] = str(datetime.now())
                failed_files['File'] = str(filename)
                failed_files['Error'] = str(err)
                table_updater('error_log','Table2',failed_files)
                if err.resp.status == 400:
                    error_details = err.content.decode('utf-8')
                    if 'API_KEY_INVALID' in error_details:
                        print("Error: Invalid API key provided.")
                        break
                    else:
                        print("Error: Bad request, but not related to API key.")
        else:
            print(f"--- SKIPPED FILE: {files.index(filename)+1} OF {len(files)}, {filename} --- NOT A VALID PDF FILE")
    else:
        print(f"--- SKIPPING FILE: {files.index(filename)+1} OF {len(files)}, {filename} --- ALREADY RECORDED")

print('******************************DONE********************************')
